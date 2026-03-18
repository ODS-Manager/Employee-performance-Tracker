"""
Billing Service - New Version
Business logic for billing reports and calculations
Billing is done organization-wide grouped by product types with shortened team names
"""
from sqlalchemy.orm import Session
from sqlalchemy import func
from sqlalchemy.exc import IntegrityError
from datetime import datetime, date
from typing import Optional, List, Dict
from io import BytesIO
from app.models.billing import BillingReport, BillingDetail
from app.models.order import Order
from app.models.team import Team
from app.models.user import User
from app.schemas.billing import (
    BillingReportCreate,
    BillingReportResponse,
    BillingDetailResponse,
    BillingPreviewRequest,
    BillingPreviewResponse,
    BillingPreviewDetail
)
from fastapi import HTTPException

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


def get_team_short_name(team_name: str) -> str:
    """
    Convert team name to shortened version for product types
    Washington -> WA, Florida -> FL, California -> CA, etc.
    """
    team_mapping = {
        'Arizona': 'AZ',
        'California': 'CA',
        'Colorado': 'CO',
        'FIF': 'FI',
        'Florida': 'FL',
        'Georgia': 'GA',
        'GI Clearing': 'GI',
        'Michigan': 'MI',
        'National Streamline': 'NS',
        'Ohio': 'OH',
        'Oregon': 'OR',
        'Pennsylvania': 'PA',
        'Regional Streamline': 'RS',
        'SCB & PD': 'SC',
        'Texas': 'TX',
        'Utah': 'UT',
        'Vietnam Team': 'VN',
        'Washington': 'WA'
    }
    return team_mapping.get(team_name, team_name[:2].upper())


def format_product_type(team_name: str, product_type: str) -> str:
    """
    Format product type with team prefix
    Example: Washington + Full Search -> "WA Full Search"
    Note: Division is added separately in grouping key
    """
    short_name = get_team_short_name(team_name)
    # Remove any existing state prefix from product type
    product_clean = product_type.strip()
    return f"{short_name} {product_clean}"


def get_billing_reports(
    db: Session,
    org_id: Optional[int],
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    status: Optional[str] = None
) -> List[BillingReportResponse]:
    """
    Get billing reports with filters
    Shows both org-wide reports (team_id=NULL) and old team-specific reports
    If org_id is None (superadmin), returns all reports from all organizations
    """
    query = db.query(BillingReport)
    
    # Filter by org_id only if provided (for superadmin, org_id might be None)
    if org_id is not None:
        query = query.filter(BillingReport.org_id == org_id)
    
    if start_date:
        query = query.filter(BillingReport.start_date >= start_date)
    if end_date:
        query = query.filter(BillingReport.end_date <= end_date)
    if status:
        query = query.filter(BillingReport.status == status)
    
    reports = query.order_by(
        BillingReport.start_date.desc()
    ).all()
    
    # Convert to response models with details
    result = []
    for report in reports:
        details = [
            BillingDetailResponse(
                id=d.id,
                state=d.state,
                productType=d.product_type,
                divisionId=d.division_id,
                divisionName="Direct" if d.division_id == 1 else "Agency",
                singleSeatCount=d.single_seat_count,
                onlyStep1Count=d.only_step1_count,
                onlyStep2Count=d.only_step2_count,
                totalCount=d.total_count
            )
            for d in report.details
        ]
        
        total_files = sum(d.total_count for d in report.details)
        
        response = BillingReportResponse(
            id=report.id,
            orgId=report.org_id,
            teamId=report.team_id,
            teamName=report.team.name if report.team else "All Teams",
            startDate=report.start_date,
            endDate=report.end_date,
            status=report.status,
            createdBy=report.created_by,
            createdByName=report.created_by_user.user_name if report.created_by_user else None,
            finalizedBy=report.finalized_by,
            finalizedByName=report.finalized_by_user.user_name if report.finalized_by_user else None,
            finalizedAt=report.finalized_at,
            createdAt=report.created_at,
            modifiedAt=report.modified_at,
            details=details,
            totalFiles=total_files
        )
        result.append(response)
    
    return result


def preview_billing_data(
    db: Session,
    org_id: int,
    request: BillingPreviewRequest
) -> BillingPreviewResponse:
    """
    Preview billing data before generating report
    Shows what will be included in the billing report
    Organization-wide, grouped by product type + division (team + product + division)
    """
    # Use the date range from request
    start_date = request.start_date
    end_date = request.end_date
    
    # Get all active teams in the organization (eager-load their products from team_products table)
    teams = db.query(Team).filter(
        Team.org_id == org_id,
        Team.is_active == True
    ).order_by(Team.name).all()
    
    # Query ALL pending orders with Completed status in the organization for this period
    # ONLY include Completed orders (order_status_id = 1) for billing
    orders = db.query(Order).filter(
        Order.org_id == org_id,
        Order.billing_status == 'pending',
        Order.order_status_id == 1,  # Only "Completed" status
        Order.entry_date >= start_date,
        Order.entry_date < end_date,
        Order.deleted_at.is_(None)
    ).all()
    
    # Initialize billing_data using each team's configured products from team_products table
    # Key format: "product_type|division_id"
    billing_data: Dict[str, Dict[str, int]] = {}
    
    # Build a team_id -> team lookup for order processing
    team_lookup: Dict[int, Team] = {team.id: team for team in teams}
    
    for team in teams:
        # Use the team's configured products from the team_products table (source of truth)
        for team_product in team.products:
            formatted_product = format_product_type(team.name, team_product.product_type)
            # Create entries for both Direct (1) and Agency (2) divisions
            for div_id in [1, 2]:  # Assuming 1=Direct, 2=Agency
                key = f"{formatted_product}|{div_id}"
                billing_data[key] = {
                    'single_seat': 0,
                    'only_step1': 0,
                    'only_step2': 0,
                    'team_name': team.name,
                    'product_type': team_product.product_type,
                    'division_id': div_id
                }
    
    # Now count actual orders
    for order in orders:
        # Get team from lookup (avoid extra DB queries)
        team = team_lookup.get(order.team_id)
        if not team:
            continue
        
        # Format product type with team prefix
        formatted_product = format_product_type(team.name, order.product_type)
        key = f"{formatted_product}|{order.division_id}"
        
        if key not in billing_data:
            # Order has a product type not configured in team_products — still include it
            billing_data[key] = {
                'single_seat': 0,
                'only_step1': 0,
                'only_step2': 0,
                'team_name': team.name,
                'product_type': order.product_type,
                'division_id': order.division_id
            }
        
        # Determine order type
        if order.step1_user_id and order.step2_user_id:
            billing_data[key]['single_seat'] += 1
        elif order.step1_user_id:
            billing_data[key]['only_step1'] += 1
        elif order.step2_user_id:
            billing_data[key]['only_step2'] += 1
    
    # Convert to preview details (include all entries, even with 0 counts)
    details = []
    for product_key, counts in sorted(billing_data.items()):
        total = counts['single_seat'] + counts['only_step1'] + counts['only_step2']
        division_name = "Direct" if counts['division_id'] == 1 else "Agency"
        details.append(
            BillingPreviewDetail(
                productType=counts['product_type'],
                divisionId=counts['division_id'],
                divisionName=division_name,
                singleSeatCount=counts['single_seat'],
                onlyStep1Count=counts['only_step1'],
                onlyStep2Count=counts['only_step2'],
                totalCount=total
            )
        )
    
    total_files = sum(d.total_count for d in details)
    teams_count = len(teams)
    
    return BillingPreviewResponse(
        startDate=start_date,
        endDate=end_date,
        details=details,
        totalFiles=total_files,
        pendingOrdersCount=len(orders),
        teamsCount=teams_count
    )


def create_billing_report(
    db: Session,
    org_id: int,
    current_user_id: int,
    data: BillingReportCreate
) -> BillingReportResponse:
    """
    Create a new organization-wide billing report grouped by product types and divisions
    """
    # Check if report already exists for this period
    existing = db.query(BillingReport).filter(
        BillingReport.org_id == org_id,
        BillingReport.team_id.is_(None),  # Org-wide reports have null team_id
        BillingReport.start_date == data.start_date,
        BillingReport.end_date == data.end_date
    ).first()
    
    if existing:
        raise HTTPException(
            status_code=400,
            detail=f"Billing report already exists for {data.start_date} to {data.end_date}. Report ID: {existing.id}"
        )
    
    # Use the date range from request
    start_date = data.start_date
    end_date = data.end_date
    
    # Get all active teams in the organization (eager-load their products from team_products table)
    teams = db.query(Team).filter(
        Team.org_id == org_id,
        Team.is_active == True
    ).order_by(Team.name).all()
    
    # Query pending orders with Completed status for this period
    # ONLY include Completed orders (order_status_id = 1) for billing
    orders = db.query(Order).filter(
        Order.org_id == org_id,
        Order.billing_status == 'pending',
        Order.order_status_id == 1,  # Only "Completed" status
        Order.entry_date >= start_date,
        Order.entry_date < end_date,
        Order.deleted_at.is_(None)
    ).all()
    
    # Create billing report (team_id is NULL for org-wide)
    report = BillingReport(
        org_id=org_id,
        team_id=None,  # Org-wide report
        start_date=start_date,
        end_date=end_date,
        status='draft',
        created_by=current_user_id
    )
    db.add(report)
    db.flush()  # Get report.id
    
    # Initialize billing_data using each team's configured products from team_products table
    # Key format: "product_type|division_id"
    billing_data: Dict[str, Dict[str, int]] = {}
    
    # Build a team_id -> team lookup for order processing
    team_lookup: Dict[int, Team] = {team.id: team for team in teams}
    
    for team in teams:
        # Use the team's configured products from the team_products table (source of truth)
        for team_product in team.products:
            formatted_product = format_product_type(team.name, team_product.product_type)
            # Create entries for both Direct (1) and Agency (2) divisions
            for div_id in [1, 2]:  # Assuming 1=Direct, 2=Agency
                key = f"{formatted_product}|{div_id}"
                billing_data[key] = {
                    'single_seat': 0,
                    'only_step1': 0,
                    'only_step2': 0,
                    'team_name': team.name,
                    'product_type': team_product.product_type,
                    'division_id': div_id
                }
    
    # Now count actual orders
    for order in orders:
        # Get team from lookup (avoid extra DB queries)
        team = team_lookup.get(order.team_id)
        if not team:
            continue
        
        # Format product type with team prefix
        formatted_product = format_product_type(team.name, order.product_type)
        key = f"{formatted_product}|{order.division_id}"
        
        if key not in billing_data:
            # Order has a product type not configured in team_products — still include it
            billing_data[key] = {
                'single_seat': 0,
                'only_step1': 0,
                'only_step2': 0,
                'team_name': team.name,
                'product_type': order.product_type,
                'division_id': order.division_id
            }
        
        # Determine order type
        if order.step1_user_id and order.step2_user_id:
            billing_data[key]['single_seat'] += 1
        elif order.step1_user_id:
            billing_data[key]['only_step1'] += 1
        elif order.step2_user_id:
            billing_data[key]['only_step2'] += 1
    
    # Create billing details for ALL combinations (including zeros)
    for product_key, counts in billing_data.items():
        total = counts['single_seat'] + counts['only_step1'] + counts['only_step2']
        detail = BillingDetail(
            report_id=report.id,
            state=None,  # NULL for org-wide reports
            product_type=counts['product_type'],
            division_id=counts['division_id'],
            single_seat_count=counts['single_seat'],
            only_step1_count=counts['only_step1'],
            only_step2_count=counts['only_step2'],
            total_count=total
        )
        db.add(detail)
    
    try:
        db.commit()
        db.refresh(report)
    except IntegrityError as e:
        db.rollback()
        # Check if it's a duplicate key error
        if 'idx_billing_org_team_period' in str(e) or 'unique constraint' in str(e).lower():
            raise HTTPException(
                status_code=400,
                detail=f"Billing report already exists for {data.start_date} to {data.end_date}. Please delete the existing report before creating a new one."
            )
        else:
            # Re-raise other integrity errors
            raise HTTPException(
                status_code=500,
                detail=f"Database error while creating billing report: {str(e)}"
            )
    
    # Return response
    return get_billing_report_by_id(db, report.id)


def get_billing_report_by_id(db: Session, report_id: int) -> BillingReportResponse:
    """Get a single billing report by ID"""
    report = db.query(BillingReport).filter(BillingReport.id == report_id).first()
    
    if not report:
        raise HTTPException(status_code=404, detail="Billing report not found")
    
    details = [
        BillingDetailResponse(
            id=d.id,
            state=d.state,
            productType=d.product_type,
            divisionId=d.division_id,
            divisionName="Direct" if d.division_id == 1 else "Agency",
            singleSeatCount=d.single_seat_count,
            onlyStep1Count=d.only_step1_count,
            onlyStep2Count=d.only_step2_count,
            totalCount=d.total_count
        )
        for d in report.details
    ]
    
    total_files = sum(d.total_count for d in report.details)
    
    return BillingReportResponse(
        id=report.id,
        org_id=report.org_id,
        team_id=None,
        team_name="All Teams",
        start_date=report.start_date,
        end_date=report.end_date,
        status=report.status,
        created_by=report.created_by,
        created_by_name=report.created_by_user.user_name if report.created_by_user else None,
        finalized_by=report.finalized_by,
        finalized_by_name=report.finalized_by_user.user_name if report.finalized_by_user else None,
        finalized_at=report.finalized_at,
        created_at=report.created_at,
        modified_at=report.modified_at,
        details=details,
        total_files=total_files
    )


def finalize_billing_report(
    db: Session,
    report_id: int,
    current_user_id: int
) -> BillingReportResponse:
    """
    Finalize a billing report
    - Marks the report as 'finalized'
    - Updates all associated orders billing_status from 'pending' to 'done'
    """
    report = db.query(BillingReport).filter(BillingReport.id == report_id).first()
    
    if not report:
        raise HTTPException(status_code=404, detail="Billing report not found")
    
    if report.status == 'finalized':
        raise HTTPException(status_code=400, detail="Report is already finalized")
    
    # Use the date range from the report
    start_date = report.start_date
    end_date = report.end_date
    
    # Update all pending Completed orders for this organization and period to 'done'
    # ONLY update Completed orders (order_status_id = 1)
    updated_count = db.query(Order).filter(
        Order.org_id == report.org_id,
        Order.billing_status == 'pending',
        Order.order_status_id == 1,  # Only "Completed" status
        Order.entry_date >= start_date,
        Order.entry_date < end_date,
        Order.deleted_at.is_(None)
    ).update({
        Order.billing_status: 'done',
        Order.modified_at: datetime.now(),
        Order.modified_by: current_user_id
    }, synchronize_session=False)
    
    # Update report status
    report.status = 'finalized'
    report.finalized_by = current_user_id
    report.finalized_at = datetime.now()
    report.modified_at = datetime.now()
    
    db.commit()
    db.refresh(report)
    
    return get_billing_report_by_id(db, report.id)


def delete_billing_report(db: Session, report_id: int) -> None:
    """
    Delete a billing report (only if in draft status)
    """
    report = db.query(BillingReport).filter(BillingReport.id == report_id).first()
    
    if not report:
        raise HTTPException(status_code=404, detail="Billing report not found")
    
    if report.status == 'finalized':
        raise HTTPException(
            status_code=400,
            detail="Cannot delete finalized billing report"
        )
    
    db.delete(report)
    db.commit()


def export_billing_report_to_excel(db: Session, report_id: int) -> BytesIO:
    """
    Export billing report to Excel format - Team x Product Type matrix
    Shows all teams and all product types (even with 0 counts)
    """
    if not EXCEL_AVAILABLE:
        raise HTTPException(
            status_code=500,
            detail="Excel export not available. Please install openpyxl package."
        )
    
    # Get report
    report = db.query(BillingReport).filter(BillingReport.id == report_id).first()
    
    if not report:
        raise HTTPException(status_code=404, detail="Billing report not found")
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Billing Report"
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Group details by team, product type, and division
    # Extract team and product from formatted product_type (e.g., "WA Full Search" -> team="Washington", product="Full Search")
    team_product_data = {}
    
    # Build reverse mapping from short code to full team name using the org's actual teams
    # This avoids any hardcoded team_name_map
    org_teams = db.query(Team).filter(
        Team.org_id == report.org_id
    ).all()
    
    # Build short_code -> team_name mapping dynamically from actual teams
    team_name_map = {}
    for t in org_teams:
        short = get_team_short_name(t.name)
        team_name_map[short] = t.name
    
    for detail in report.details:
        # Extract team code and product from product_type
        parts = detail.product_type.split(' ', 1)
        if len(parts) >= 2:
            team_code = parts[0]
            product = parts[1].strip()
            team_name = team_name_map.get(team_code, team_code)
            
            if team_name not in team_product_data:
                team_product_data[team_name] = {}
            
            # Initialize product entry if not exists
            if product not in team_product_data[team_name]:
                team_product_data[team_name][product] = {
                    'direct_single_seat': 0,
                    'direct_step1': 0,
                    'direct_step2': 0,
                    'agency_single_seat': 0,
                    'agency_step1': 0,
                    'agency_step2': 0,
                    'single_seat': 0,  # For totals
                    'step1': 0,        # For totals
                    'step2': 0,        # For totals
                    'total': 0
                }
            
            # Distribute counts to appropriate division
            division_prefix = 'direct_' if detail.division_id == 1 else 'agency_'
            team_product_data[team_name][product][division_prefix + 'single_seat'] = detail.single_seat_count
            team_product_data[team_name][product][division_prefix + 'step1'] = detail.only_step1_count
            team_product_data[team_name][product][division_prefix + 'step2'] = detail.only_step2_count
            
            # Update totals
            team_product_data[team_name][product]['single_seat'] += detail.single_seat_count
            team_product_data[team_name][product]['step1'] += detail.only_step1_count
            team_product_data[team_name][product]['step2'] += detail.only_step2_count
            team_product_data[team_name][product]['total'] += detail.total_count
    
    # Get all unique product types and teams
    all_products = sorted(set(
        product 
        for products in team_product_data.values() 
        for product in products.keys()
    ))
    
    # Sort teams by org team order (as they appear in the database), then alphabetically for any extras
    team_order = [t.name for t in org_teams]
    
    # Sort teams by the defined order
    sorted_teams = sorted(
        team_product_data.keys(),
        key=lambda t: team_order.index(t) if t in team_order else 999
    )
    
    # Header row
    header_row = 1
    headers = ["Team Name", "Product Type", "Division", "Process Type", "Count"]
    for col_idx, header_text in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = header_text
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Write data rows - 6 sub-rows per product type (3 process types x 2 divisions)
    current_row = header_row + 1
    grand_total_single_seat = 0
    grand_total_step1 = 0
    grand_total_step2 = 0
    
    for team_name in sorted_teams:
        products = team_product_data[team_name]
        sorted_products = sorted(products.keys())
        team_first_row = True
        
        for product in sorted_products:
            counts = products[product]
            grand_total_single_seat += counts['single_seat']
            grand_total_step1 += counts['step1']
            grand_total_step2 += counts['step2']
            
            sub_rows = [
                ('Direct', 'Single Seat', counts['direct_single_seat']),
                ('Direct', 'Step 1', counts['direct_step1']),
                ('Direct', 'Step 2', counts['direct_step2']),
                ('Agency', 'Single Seat', counts['agency_single_seat']),
                ('Agency', 'Step 1', counts['agency_step1']),
                ('Agency', 'Step 2', counts['agency_step2']),
            ]
            
            for sub_idx, (division_label, process_label, count) in enumerate(sub_rows):
                # Team name only on first sub-row of the first product of each team
                if team_first_row and sub_idx == 0:
                    ws.cell(row=current_row, column=1).value = team_name
                    ws.cell(row=current_row, column=1).font = Font(bold=True)
                    team_first_row = False
                else:
                    ws.cell(row=current_row, column=1).value = ""
                
                # Product name only on first sub-row of each product
                if sub_idx == 0:
                    ws.cell(row=current_row, column=2).value = product
                    ws.cell(row=current_row, column=2).font = Font(bold=False)
                else:
                    ws.cell(row=current_row, column=2).value = ""
                
                ws.cell(row=current_row, column=3).value = division_label
                ws.cell(row=current_row, column=4).value = process_label
                ws.cell(row=current_row, column=5).value = count
                
                # Apply borders and alignment
                for col in range(1, 6):
                    cell = ws.cell(row=current_row, column=col)
                    cell.border = border
                    if col == 5:  # Count column
                        cell.alignment = center_align
                
                current_row += 1
    
    # Grand total rows
    total_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Single Seat total
    ws.cell(row=current_row, column=1).value = "GRAND TOTAL"
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    ws.cell(row=current_row, column=2).value = ""
    ws.cell(row=current_row, column=3).value = ""
    ws.cell(row=current_row, column=4).value = "Single Seat"
    ws.cell(row=current_row, column=4).font = Font(bold=True)
    ws.cell(row=current_row, column=5).value = grand_total_single_seat
    for col in range(1, 6):
        cell = ws.cell(row=current_row, column=col)
        cell.fill = total_fill
        cell.font = Font(bold=True)
        cell.border = border
        if col == 5:
            cell.alignment = center_align
    current_row += 1
    
    # Step 1 total
    ws.cell(row=current_row, column=1).value = ""
    ws.cell(row=current_row, column=2).value = ""
    ws.cell(row=current_row, column=3).value = ""
    ws.cell(row=current_row, column=4).value = "Step 1"
    ws.cell(row=current_row, column=4).font = Font(bold=True)
    ws.cell(row=current_row, column=5).value = grand_total_step1
    for col in range(1, 6):
        cell = ws.cell(row=current_row, column=col)
        cell.fill = total_fill
        cell.font = Font(bold=True)
        cell.border = border
        if col == 5:
            cell.alignment = center_align
    current_row += 1
    
    # Step 2 total
    ws.cell(row=current_row, column=1).value = ""
    ws.cell(row=current_row, column=2).value = ""
    ws.cell(row=current_row, column=3).value = ""
    ws.cell(row=current_row, column=4).value = "Step 2"
    ws.cell(row=current_row, column=4).font = Font(bold=True)
    ws.cell(row=current_row, column=5).value = grand_total_step2
    for col in range(1, 6):
        cell = ws.cell(row=current_row, column=col)
        cell.fill = total_fill
        cell.font = Font(bold=True)
        cell.border = border
        if col == 5:
            cell.alignment = center_align
    current_row += 1
    
    # Overall total row
    total_files = grand_total_single_seat + grand_total_step1 + grand_total_step2
    overall_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
    ws.cell(row=current_row, column=1).value = "TOTAL"
    ws.cell(row=current_row, column=2).value = ""
    ws.cell(row=current_row, column=3).value = ""
    ws.cell(row=current_row, column=4).value = ""
    ws.cell(row=current_row, column=5).value = total_files
    for col in range(1, 6):
        cell = ws.cell(row=current_row, column=col)
        cell.fill = overall_fill
        cell.font = Font(bold=True, size=12)
        cell.border = border
        if col == 5:
            cell.alignment = center_align
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25  # Team Name
    ws.column_dimensions['B'].width = 25  # Product Type
    ws.column_dimensions['C'].width = 15  # Division
    ws.column_dimensions['D'].width = 15  # Process Type
    ws.column_dimensions['E'].width = 12  # Count
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file
